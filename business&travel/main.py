# -*- coding: utf-8 -*-
"""
@author: jmartorell
"""
import time
import operator
from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Side, Border
from openpyxl import load_workbook
from openpyxl import Workbook
import targetX as t
import dataset as d
import os


class App:
    # 1) init method initializes variables that will be accessible by self from any method of the class
    def __init__(self, username='business.travel', password='Busi2016', target_destination='New york', depart_m='2', depart_w='3',
                 depart_d='1', return_m='2', return_w='3', return_d='7', cell_city='New York', cell_cc='US',
                 path='/home/jmartorell/Booking'):
        self.username = username
        self.password = password
        self.target_destination = target_destination
        self.depart_m = depart_m
        self.depart_w = depart_w
        self.depart_d = depart_d
        self.return_m = return_m
        self.return_w = return_w
        self.return_d = return_d
        self.cell_city = cell_city
        self.cell_cc = cell_cc
        self.path = path
        self.browser = webdriver.Firefox(
            executable_path='/usr/local/bin/geckodriver')
        self.error = False
        self.url = 'https://pro.w2m.travel'
        self.all_hotels = []
        self.all_prices = []
        self.all_locations = []
        self.all_positions = []
        self.display = []
        self.cheap = []
        self.data = {}
        self.index = ""
        self.euro_symbol = '€'
        self.coefficient = '1.374'
        self.browser.get(self.url)
        self.log_in()
        if self.error is False:
            self.search_engine_insert()
        if self.error is False:
            self.scroll_down()
        if self.error is False:
            if not os.path.exists(path):
                os.mkdir(path)
            self.file_manager()
            # todo: self.reach_target()
        # close the browser
        sleep(1)
        self.browser.quit()

    # 2) log in method allows us to log in to access the provider's services
    def log_in(self, ):
        try:
            input_element = self.browser.find_element_by_id("email")
            input_element.clear()
            input_element.send_keys("business.travel")
            input_element = self.browser.find_element_by_id("password")
            input_element.clear()
            print('Logging in with username and password ...')
            input_element.send_keys("Busi2016")
            input_element.submit()

            # print(self.browser.current_url)
        except Exception:
            print('Some exception occurred while trying to find username or password field')
            self.error = True

    def cookies_popup(self):
        print('closing cookies window ...')
        WebDriverWait(self.browser, 100).until(EC.visibility_of_element_located((
            By.CSS_SELECTOR,
            '.cookie-policy__close'))).click()
        sleep(1)

    # 3) search engine insert method fills in the search engine fields and clicks on the search button
    def search_engine_insert(self):
        # wait to load the search engine
        WebDriverWait(self.browser, 100).until(EC.visibility_of_element_located((
            By.XPATH,
            '//*[@id="hotel-searcher-_ctl1__ctl1__ctl1_pageBody_pageBody_searcher__ctl0_ctlZoneSelector-input"]'))).click()
        self.cookies_popup()
        element = self.browser.find_element_by_xpath(
            '//*[@id="hotel-searcher-_ctl1__ctl1__ctl1_pageBody_pageBody_searcher__ctl0_ctlZoneSelector-input"]')
        element.clear()

        # check access
        # assert "Hoteles | W2M" in self.browser.title

        # enter data in input field
        element.send_keys(self.target_destination)

        # TODO:
        # drop-down item selection
        actions = ActionChains(self.browser)
        for _ in range(1):
            actions.send_keys(Keys.ARROW_DOWN).perform()
            sleep(1)

        # enter destination city
        target_city = element.find_element_by_xpath(
            "//div[3]/div[1]")
        target_city.click()

        # press the search button
        login_attempt = element.find_element_by_xpath(
            "//div[2]/div[2]/button")
        print('loading page ...')
        login_attempt.click()

    # 4) the reach target method systematically selects the first objective by clicking on it
    # todo: def reach_target(self):
    #     element = WebDriverWait(self.browser, 10).until(EC.visibility_of_element_located((
    #         By.XPATH,
    #         '//main/div[1]/div/div[1]/article/div[1]/div[2]/h2'))).click()
    def reach_target(self, index):
        target_button = self.browser.find_element_by_xpath(
            '//div[ ' + index + ' ]/article/div[1]/div[2]/div[2]/div/div[2]/span/a')
        self.browser.execute_script("arguments[0].scrollIntoView();", target_button)
        # target_button.click()

    def scroll_down(self):
        self.browser.implicitly_wait(15)

        # todo REF: https://stackoverflow.com/questions/48006078/how-to-scroll-down-in-python-selenium-step-by-step
        read_mores = self.browser.find_elements_by_xpath('//div[text()="Best stay price"]')
        print('Scrolling page ...')
        for read_more in read_mores:
            self.browser.execute_script("arguments[0].scrollIntoView();", read_more)
            # read_more.click()

        print("Scraping page ...")
        soup = BeautifulSoup(self.browser.page_source, 'lxml')
        hotel_list = soup.find_all('div', {'class': 'results-list__item'})

        print("\n\tdisplay:\n")
        try:
            for i, hotel in enumerate(hotel_list):
                self.all_positions.append(i + 1)
                hotel_name = hotel.find('h2', {'class': 'info-card__title'}).getText()
                # fixme: remove whitespaces REF: https://stackoverrun.com/es/q/743639
                hotel_name = ' '.join(hotel_name.split())
                # notice that instead of .getText().strip('€') here we work with .getText().replace('€', ''):
                hotel_price = hotel.find('a', {'tabindex': '0'}).getText().replace('€', '')
                hotel_price = hotel_price.replace('.', '')
                hotel_price = hotel_price.replace(',', '.')
                hotel_price = float(hotel_price)
                hotel_price = "{0:.2f}".format(hotel_price)
                self.all_prices.append(hotel_price)

                hotel_location = hotel.find('div', {'class': 'info-card__location'}).getText().strip(',')
                hotel_location = ' '.join(hotel_location.split()).rstrip("View the map")
                self.all_locations.append(hotel_location)

                if len(hotel_price) == 5:
                    hotel_price = "   " + hotel_price
                if len(hotel_price) == 6:
                    hotel_price = "  " + hotel_price
                if len(hotel_price) == 7:
                    hotel_price = " " + hotel_price
                if len(hotel_price) == 8:
                    hotel_price = "" + hotel_price
                self.all_hotels.append(hotel_name)
                if i < 9:
                    print(" %d - %s %s %s - %s" % (i + 1, hotel_price, self.euro_symbol, hotel_name, hotel_location))
                else:
                    print("%d - %s %s %s - %s" % (i + 1, hotel_price, self.euro_symbol, hotel_name, hotel_location))

            print("\n\tranking:\n")
            # float cast
            new_prices = []
            for element in self.all_prices:
                rank = float(element)
                new_prices.append(rank)

            # final list
            display_list = zip(self.all_positions, self.all_hotels, new_prices, self.all_locations)
            ranking = sorted(display_list, key=operator.itemgetter(2))
            for j, k, v, w in ranking:
                if v < 100.00:
                    print("   ", "{0:.2f}".format(v), k)
                if 99.00 < v < 1000.00:
                    print("  ", "{0:.2f}".format(v), k)
                if 999.00 < v < 10000.00:
                    print(" ", "{0:.2f}".format(v), k)
                if v > 9999.00:
                    print("", "{0:.2f}".format(v), k)

            self.display = display_list
            self.data = ranking
            self.cheap = ranking[0]
            print('\nlocated booking for', self.cheap[2], self.euro_symbol, '...')
            print('pointing to the target button', self.cheap[0], '...')
            self.index = str(self.cheap[0])
            if self.error is False:
                self.reach_target(self.index)

            sleep(2)
        except Exception as e:
            self.error = True
            print(e)
            print('Some error occurred while trying to scroll down')

    def set_stylesheet(self, sheet):

        # time frame:
        sheet.merge_cells('A1:L1')
        time_frame = sheet['A1']
        time_frame.fill = PatternFill(
            start_color="0007147A", end_color="0007147A", fill_type="solid")
        time_frame.font = Font(bold=True, size=11)
        bd = Side(style='thick', color="000000")
        time_frame.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        # timestamp
        time_label = 'Business&Travel:        %s                       Time Frame:        %s%s/2020  -  %s%s/2020' \
                     % (time.ctime(), t.dep + "/", t.start_month, t.ret + "/", t.end_month)
        sheet.cell(row=1, column=1).value = time_label

        # title bar
        # fixme REF:
        # https://stackoverflow.com/questions/35918504/adding-a-background-color-to-cell-openpyxl
        for col_range in range(1, 13):
            cell_title = sheet.cell(2, col_range)
            cell_title.fill = PatternFill(start_color="00f4f4f7", end_color="00f4f4f7", fill_type="solid")
            cell_title = sheet.cell(2, col_range)
            cell_title.font = Font(bold=True, size=11)
            bd = Side(style='thick', color="000000")
            cell_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        header = ('Code', 'Price', 'Retail', 'Profit', 'CC', 'City', 'No', 'Hotel', 'Gr', 'Co', 'Location', 'Density')
        sheet.cell(row=2, column=1).value = header[0]
        sheet.cell(row=2, column=2).value = header[1]
        sheet.cell(row=2, column=3).value = header[2]
        sheet.cell(row=2, column=4).value = header[3]
        sheet.cell(row=2, column=5).value = header[4]
        sheet.cell(row=2, column=6).value = header[5]
        sheet.cell(row=2, column=7).value = header[6]
        sheet.cell(row=2, column=8).value = header[7]
        sheet.cell(row=2, column=9).value = header[8]
        sheet.cell(row=2, column=10).value = header[9]
        sheet.cell(row=2, column=11).value = header[10]
        sheet.cell(row=2, column=12).value = header[11]

        sheet.column_dimensions['B'].number_format = '#,##0.00'
        sheet.column_dimensions['C'].number_format = '#,##0.00'
        sheet.column_dimensions['D'].number_format = '#,##0.00'
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 9
        sheet.column_dimensions['C'].width = 9
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 4
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 4
        sheet.column_dimensions['H'].width = 60
        sheet.column_dimensions['I'].width = 4
        sheet.column_dimensions['J'].width = 4
        sheet.column_dimensions['K'].width = 50
        sheet.column_dimensions['L'].width = 18

        format = sheet.column_dimensions['A']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['B']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['C']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['D']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['E']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['F']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['G']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['H']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['I']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['J']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['K']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['L']
        format.font = Font(bold=True, italic=True, name='Arial')

    def read_code(self):
        global trip_code
        f = open("trip_code.txt", "r")
        if f.mode == 'r':
            trip_code = f.read()
        return trip_code

    def write_code(self, input_code):
        f = open("trip_code.txt", "w")
        f.write(input_code)
        f.close()

    def write_bookings_to_excel_file(self, booking_path):

        # FIXME: openpyxl -> https://openpyxl.readthedocs.io/en/stable/index.html
        filepath = os.path.join(booking_path, 'bookings.xlsx')
        print('Writing to excel ...')
        if not os.path.exists(filepath):
            workbook = Workbook()
            workbook.save(filepath)
            workbook.create_sheet("Spapshoot", 0)
            workbook.create_sheet("Display", 1)
        else:
            workbook = load_workbook(filepath)

        # fixme: delete the default sheet:
        if "Sheet" in workbook.sheetnames:
            std = workbook["Sheet"]
            workbook.remove(std)

        # switch sheet
        workbook.active = 0

        sheet = workbook.active
        self.set_stylesheet(sheet)

        # write sheet
        i = 3
        for row in self.data:
            cell_reference = sheet.cell(row=i, column=1)
            update_code = t.code_builder(self.read_code())
            self.write_code(update_code)
            cell_reference.value = update_code
            cell_reference = sheet.cell(row=i, column=2)
            cell_reference.value = row[2]
            sheet['C{}'.format(i)] = '=PRODUCT(B{},{}'.format(i, self.coefficient)
            sheet['D{}'.format(i)] = '=SUM(C{},-B{}'.format(i, i)
            cell_reference = sheet.cell(row=i, column=5)
            cell_reference.value = self.cell_cc
            cell_reference = sheet.cell(row=i, column=6)
            cell_reference.value = self.cell_city
            cell_reference = sheet.cell(row=i, column=7)
            cell_reference.value = row[0]
            cell_reference = sheet.cell(row=i, column=8)
            cell_reference.value = row[1]
            # cell_reference = sheet.cell(row=i, column=6)
            # cell_reference.value = self.cell_gr
            # cell_reference = sheet.cell(row=i, column=6)
            # cell_reference.value = self.cell_co
            cell_reference = sheet.cell(row=i, column=11)
            cell_reference.value = row[3]
            # cell_reference = sheet.cell(row=i, column=6)
            # cell_reference.value = self.cell_density
            i += 1

        workbook.save(filepath)  # save file

    def file_manager(self, ):
        f = open("trip_code.txt", "w+")
        f.write("LM30")
        f.close()
        bookings_folder_path = os.path.join(self.path, 'bookings')
        if not os.path.exists(bookings_folder_path):
            os.mkdir(bookings_folder_path)
        if self.error is False:
            self.write_bookings_to_excel_file(bookings_folder_path)


if __name__ == '__main__':
    switch = t.switch
    if switch != 0:
        x = 0
        while x < 24:
            app = App(depart_m=t.depart_month,
                      depart_w=t.depart_week,
                      depart_d=t.depart_day,
                      return_m=t.return_month,
                      return_w=t.return_week,
                      return_d=t.return_day,
                      target_destination=d.tour_en[x][0],
                      cell_city=d.tour_en[x][0],
                      cell_cc=d.tour_en[x][1]
                      )
            x += 1
    else:
        app = App(depart_m=t.depart_month,
                  depart_w=t.depart_week,
                  depart_d=t.depart_day,
                  return_m=t.return_month,
                  return_w=t.return_week,
                  return_d=t.return_day,
                  )
