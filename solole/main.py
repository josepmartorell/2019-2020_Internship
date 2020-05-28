import time
import operator
from time import sleep
from email import encoders
from bs4 import BeautifulSoup
from selenium import webdriver
# todo: -> Check driver version periodically
# check = input("\nPress enter to check Selenium driver version...")
# os.system('python -c "import selenium; print(selenium.__version__)"')
# As you are using Selenium 3.8.0 you have to use GeckoDriver mandatory. But again as you are using Firefox v46.0 you
# have to set the capability marionette to False through DesiredCapabilities() as follows REF:
# https://stackoverflow.com/questions/47782650/selenium-common-exceptions-sessionnotcreatedexception-message-unable-to-find-a/47785513
# from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
# cap = DesiredCapabilities().FIREFOX
# cap["marionette"] = False
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.styles import Alignment, Protection
from openpyxl import load_workbook
from openpyxl import Workbook
from os import remove
import target as t
import data as d
import datetime
import requests
import smtplib
import shutil
import ssl
import os


class App:

    def __init__(self, username='BUSINESSTRAVEL', password='Trav567RT', target_city='new york', depart_m='2', depart_w='3',
                 depart_d='1', return_m='2', return_w='3', return_d='7', cell_city='New York', cell_cc='US',
                 path='/home/jmartorell/Booking'):  # Change this to your Target details and desired booking path
        self.username = username
        self.password = password
        self.target_city = target_city
        self.depart_m = depart_m
        self.depart_w = depart_w
        self.depart_d = depart_d
        self.return_m = return_m
        self.return_w = return_w
        self.return_d = return_d
        self.cell_city = cell_city
        self.cell_cc = cell_cc
        self.path = path
        # self.driver = webdriver.Firefox(capabilities=cap, executable_path='/usr/local/bin/geckodriver')  # Change
        # this to your FirefoxDriver path. todo: the expresion "executable_path=' was missed in the original version
        #  uncapable of locating the driver!!!!
        self.driver = webdriver.Firefox(
            executable_path='/usr/local/bin/geckodriver')  # Change this to your FirefoxDriver path.
        self.error = False
        self.timeout = 30
        self.main_url = 'https://b2b.solole.es'
        self.all_positions = []
        self.all_hotels = []
        self.all_addresses = []
        self.all_prices = []
        self.euro_symbol = '€'
        self.display = []
        self.cheap = []
        self.index = ""
        self.data = {}
        self.shift = 1
        self.fork = 0
        self.switch = 1
        self.position = 0
        self.driver.get(self.main_url)
        self.log_in()
        if self.error is False:
            self.search_target_profile()
        if self.error is False:
            self.scroll_down()
        if self.error is False:
            if not os.path.exists(path):
                os.mkdir(path)
            self.file_manager()
        if self.switch != 1:
            sleep(10)
        self.driver.close()

    def log_in(self, ):
        try:
            print('\nLogging in with username and password ...')
            user_name_input = self.driver.find_element_by_xpath('//input[@placeholder="Nombre de usuario"]')
            user_name_input.send_keys(self.username)
            sleep(1)

            password_input = self.driver.find_element_by_xpath('//input[@placeholder="Contraseña"]')
            password_input.send_keys(self.password)
            sleep(1)

            password_input.submit()
            sleep(1)

            # self.close_settings_window_if_there()
        except Exception:
            print('Some exception occurred while trying to find username or password field')
            self.error = True

    def search_target_profile(self):
        try:
            print("Manipulating search engine ...")
            search_bar = self.driver.find_element_by_xpath('//*[@id="hotelzonePred"]')
            search_bar.send_keys(self.target_city)
            # target_profile_url = self.main_url + '/' + self.target_city + '/'
            # self.driver.get(target_profile_url)
            sleep(1)
            # todo: accessing a drop-down menu item directly with xpath
            # element = self.driver.find_element_by_xpath('//iboosy-hotelzone/div[2]/div/button[2]/div')
            # element.click()
            # todo: accessing a drop-down menu item by position within the list
            #  https://selenium-python.readthedocs.io/navigating.html#interacting-with-the-page
            all_options = self.driver.find_elements_by_class_name('dropdown-item')
            all_options[0].click()
            sleep(1)

            self.driver.find_element_by_css_selector(
                'div.w-50:nth-child(1) > div:nth-child(2) > div:nth-child(1)').click()
            # todo: Within <div class = "ngb-dp-week ..."> the seven days of that week are stored each in a <div
            #  class = "ngb-dp-day ...">. Secondary click on the inspector on the day that interests you and copy the
            #  css selector, which you will use to click on in the calendar picker
            self.driver.find_element_by_css_selector('div.ngb-dp-month:nth-child( ' + self.depart_m + ' ) > '
                                                     'ngb-datepicker-month-view:nth-child(1) > div:nth-child( ' + self.depart_w + ' ) > '
                                                     'div:nth-child( ' + self.depart_d + ' )').click()
            self.driver.find_element_by_css_selector('div.ngb-dp-month:nth-child( ' + self.return_m + ' ) > '
                                                     'ngb-datepicker-month-view:nth-child(1) > div:nth-child( ' + self.return_w + ' ) > '
                                                     'div:nth-child( ' + self.return_d + ' )').click()

            user_name_input = self.driver.find_element_by_xpath('//*[@id="nationalityPred"]')
            user_name_input.send_keys('España')
            sleep(1)
            # todo: accessing a drop-down menu item directly with xpath
            element = self.driver.find_element_by_xpath(
                '//div[3]/iboosy-nationalities/div/div/ngb-typeahead-window/button/div/span[2]')
            element.click()
            login_button = self.driver.find_element_by_xpath('//*[@id="searchbtn"]')
            # instead of submit it works with click
            print('Loading page ...')
            login_button.click()
            sleep(1)

        except Exception:
            self.error = True
            print('Could not find search bar')

    def scroll_down(self):
        self.driver.implicitly_wait(20)

        try:
            # self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')  # fixme: scroll
            # todo REF: https://stackoverflow.com/questions/48006078/how-to-scroll-down-in-python-selenium-step-by-step
            # FIXME 1: two ways to scroll down,
            #  1) go down to the bottom of the page at once.
            # self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')  # fixme: scroll
            # FIXME 2:
            #  2) Descend from item to item to the bottom of the page.
            # in this example and item is the text of the button "See options":
            read_mores = self.driver.find_elements_by_xpath('//div[text()="Precio desde"]')
            screen = 0
            for read_more in read_mores:
                if screen == 4:
                    print('Scrolling page ...')
                self.driver.execute_script("arguments[0].scrollIntoView();", read_more)
                screen += 1
                # read_more.click()

            try:
                soup = BeautifulSoup(self.driver.page_source, 'lxml')  # todo: bs4
                hotel_list = soup.find_all('div', {'class': 'row result-option'})
                # fixme: name mechanism:
                for i, hotel in enumerate(hotel_list):
                    self.all_positions.append(i + 1)
                    hotel_name = hotel.find('span', {'_ngcontent-c18': ""}).getText()
                    hotel_name = ' '.join(hotel_name.split())
                    # print("%d - %s" % (i + 1, hotel_name))
                    self.all_hotels.append(hotel_name)
            except IOError as e:
                print("I/O error occurred: ", os.strerror(e.errno))
                print("Error loading the hotels ")
                pass

            try:
                soup = BeautifulSoup(self.driver.page_source, 'lxml')  # todo: bs4
                address_list = soup.find_all('div', {'class': 'address'})
                # fixme: address mechanism:
                for i, address in enumerate(address_list):
                    hotel_address = address.find('span', {'_ngcontent-c18': ""}).getText()
                    hotel_address = ' '.join(hotel_address.split())
                    self.all_addresses.append(hotel_address)
                print("Scraping page ...")
            except IOError as e:
                print("I/O error occurred: ", os.strerror(e.errno))
                print("Error loading addresses ")
                pass

            try:
                soup = BeautifulSoup(self.driver.page_source, 'lxml')  # todo: bs4
                price_list = soup.find_all('div', {'class': 'text-main-light prices'})
                # fixme: price mechanism:
                for i, price in enumerate(price_list):
                    hotel_price = price.find('span', {'_ngcontent-c18': ""}).getText().replace('€', '')
                    hotel_price = ' '.join(hotel_price.split())
                    if len(hotel_price) == 5:
                        hotel_price = "   " + hotel_price
                    if len(hotel_price) == 6:
                        hotel_price = "  " + hotel_price
                    if len(hotel_price) == 7:
                        hotel_price = " " + hotel_price
                    if len(hotel_price) == 8:
                        hotel_price = "" + hotel_price
                    self.all_prices.append(hotel_price)

            except IOError as e:
                print("I/O error occurred: ", os.strerror(e.errno))
                print("Error loading prices ")
                pass

            print("\n\tSnapshoot:\n")

            # display list
            list = zip(self.all_prices, self.all_hotels, self.all_addresses)
            for i, (j, k, v) in enumerate(list):
                if len(j) == 5:
                    j = "   " + j
                if len(j) == 6:
                    j = "  " + j
                if len(j) == 7:
                    j = " " + j
                if len(j) == 8:
                    j = "" + j
                if i < 9:
                    print(" %d - %s %s %s %s %s" % (i + 1, j, self.euro_symbol, k, " - ", v))
                else:
                    print("%d - %s %s %s %s %s" % (i + 1, j, self.euro_symbol, k, " - ", v))

            print("\n\tRanking:\n")

            # float cast
            new_prices = []
            for element in self.all_prices:
                rank = float(element)
                new_prices.append(rank)

            # final list
            display_list = zip(self.all_positions, self.all_hotels, new_prices, self.all_addresses)
            ranking = sorted(display_list, key=operator.itemgetter(2))
            for j, k, v, w in ranking:
                if v < 100.00:
                    print("   ", "{0:.2f}".format(v), k)
                if 99.00 < v < 1000.00:
                    print("  ", "{0:.2f}".format(v), k)
                if 999.00 < v < 10000.00:
                    print(" ", "{0:.2f}".format(v), k)
                if v > 9999.00:
                    print("", "{0:.2f}".format(v), k)

            self.display = display_list
            self.data = ranking
            self.cheap = ranking[0]
            print('\nCheapest reservations: ', self.cheap[1], self.cheap[2], self.euro_symbol)
            for i, collation in enumerate(display_list):
                if collation[1] == self.cheap[1]:
                    self.position = i
            print('Pointing to the target button ', self.position + 1, ' ...')
            # FIXME WARNING!!!! next line does not work with position -1 or just position! see it why...
            # Coincidentally, the first is the cheapest, that is, index 1, the variable self.index starts with 0,
            # therefore we must add 1. If we subtracted  -1 or we did not add anything, it was out of range,
            # and the spider did not I found the address.
            self.index = str(self.position + 1)
            if self.error is False:
                self.target_button(self.index)

        except NoSuchElementException:
            print('Some error occurred while trying to scroll down')
            self.error = True

    def target_button(self, index):
        if index != '1':
            target_button = self.driver.find_element_by_xpath(
                '//div[ ' + index + ' ]/div/div[1]/div[2]/div[2]/div[1]/div[2]/span')
            self.driver.execute_script("arguments[0].scrollIntoView();", target_button)
            # target_button.click()
        else:
            target_button = self.driver.find_element_by_xpath(
                '//ng-component/div/div[1]/iboosy-navigation-bar/nav/div[2]/div[1]/a/div')
            self.driver.execute_script("arguments[0].scrollIntoView();", target_button)
            # target_button.click()

    def file_manager(self, ):
        f = open("trip_code.txt", "w+")
        f.write("LM30")
        f.close()
        bookings_folder_path = os.path.join(self.path, 'bookings')
        if not os.path.exists(bookings_folder_path):
            os.mkdir(bookings_folder_path)
        if self.error is False:
            self.write_bookings_to_excel_file(bookings_folder_path, self.shift)
        # if self.error is False:
        #     self.read_bookings_from_excel_file(self.path + '/bookings/bookings.xlsx')
        return self.shift

    def read_code(self):
        global trip_code
        f = open("trip_code.txt", "r")
        if f.mode == 'r':
            trip_code = f.read()
        return trip_code
    def write_code(self, input_code):
        f = open("trip_code.txt", "w")
        f.write(input_code)
        f.close()
        pass

    def set_stylesheet(self, sheet, shift):
        # snap style sheet:
        if shift == 0:
            # set title header:
            header = ('Price', 'Retail', 'Profit', 'No', 'Hotel', 'Address')
            sheet.cell(row=1, column=1).value = header[0]
            sheet.cell(row=1, column=2).value = header[1]
            sheet.cell(row=1, column=3).value = header[2]
            sheet.cell(row=1, column=4).value = header[3]
            sheet.cell(row=1, column=5).value = header[4]
            sheet.cell(row=1, column=6).value = header[5]
            # set column width:
            sheet.column_dimensions['A'].width = 9
            sheet.column_dimensions['B'].width = 9
            sheet.column_dimensions['C'].width = 9
            sheet.column_dimensions['D'].width = 4
            sheet.column_dimensions['E'].width = 60
            sheet.column_dimensions['F'].width = 50
            # fixme: set number format: (only works in white sheets)
            sheet.column_dimensions['A'].number_format = '#,##0.00'
            sheet.column_dimensions['B'].number_format = '#,##0.00'
            sheet.column_dimensions['C'].number_format = '#,##0.00'
            # set bar title style:
            for col_range in range(1, 7):
                cell_title = sheet.cell(1, col_range)
                cell_title.fill = PatternFill(
                    start_color="00c0c0c0", end_color="00c0c0c0", fill_type="solid")
                cell_title = sheet.cell(1, col_range)
                cell_title.font = Font(bold=True, size=11)
                bd = Side(style='thick', color="000000")
                cell_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            # unwrap curtain
            for raw_range in range(len(self.all_positions)):
                for col_range in range(1, 7):
                    # fixme REF: https://stackoverrun.com/es/q/3321778
                    #  run number format with styles!
                    # _cell = ws.cell('A1')
                    # _cell.number_format = '0.00E+00'
                    _cell = sheet.cell(raw_range + 2, 1)
                    _cell.number_format = '#,##0.00'
                    _cell = sheet.cell(raw_range + 2, 2)
                    _cell.number_format = '#,##0.00'
                    _cell = sheet.cell(raw_range + 2, 3)
                    _cell.number_format = '#,##0.00'
                    cell_title = sheet.cell(raw_range + 2, col_range)
                    cell_title.fill = PatternFill(
                        start_color="00eaeaea", end_color="00eaeaea", fill_type="solid")
                    cell_title = sheet.cell(raw_range + 2, col_range)
                    cell_title.font = Font(bold=True, size=11)
                    bd = Side(style='thin', color="000000")
                    cell_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        # turbo style sheet:
        else:
            # time frame:
            sheet.merge_cells('A1:I1')
            time_frame = sheet['A1']
            time_frame.fill = PatternFill(
                start_color="00FF0000", end_color="00FF0000", fill_type="solid")
            time_frame.font = Font(bold=True, size=11)
            bd = Side(style='thick', color="000000")
            time_frame.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            # timestamp
            time_label = 'Snapshoot:        %s                       Time Frame:        %s%s/2020  -  %s%s/2020' \
                         % (time.ctime(), t.dep + "/", t.start_month, t.ret + "/", t.end_month)
            sheet.cell(row=1, column=1).value = time_label
            # set title header:
            header = ('Code', 'Price', 'Retail', 'Profit', 'CC', 'City', 'No', 'Hotel', 'Address')
            sheet.cell(row=2, column=1).value = header[0]
            sheet.cell(row=2, column=2).value = header[1]
            sheet.cell(row=2, column=3).value = header[2]
            sheet.cell(row=2, column=4).value = header[3]
            sheet.cell(row=2, column=5).value = header[4]
            sheet.cell(row=2, column=6).value = header[5]
            sheet.cell(row=2, column=7).value = header[6]
            sheet.cell(row=2, column=8).value = header[7]
            sheet.cell(row=2, column=9).value = header[8]
            # set number format:
            # fixme: set number format: (only for in white sheets)
            sheet.column_dimensions['B'].number_format = '#,##0.00'
            sheet.column_dimensions['C'].number_format = '#,##0.00'
            sheet.column_dimensions['D'].number_format = '#,##0.00'
            # set column width:
            sheet.column_dimensions['A'].width = 6
            sheet.column_dimensions['B'].width = 9
            sheet.column_dimensions['C'].width = 9
            sheet.column_dimensions['D'].width = 9
            sheet.column_dimensions['E'].width = 4
            sheet.column_dimensions['F'].width = 16
            sheet.column_dimensions['G'].width = 4
            sheet.column_dimensions['H'].width = 60
            sheet.column_dimensions['I'].width = 50
            # set bar title style:
            for col_range in range(1, 10):
                cell_title = sheet.cell(2, col_range)
                cell_title.fill = PatternFill(
                    start_color="00c0c0c0", end_color="00c0c0c0", fill_type="solid")
                cell_title = sheet.cell(2, col_range)
                cell_title.font = Font(bold=True, size=11)
                bd = Side(style='thick', color="000000")
                cell_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            # unwrap curtain
            for raw_range in range(len(self.all_positions)):
                for col_range in range(1, 10):
                    _cell = sheet.cell(raw_range + 3, 2)
                    _cell.number_format = '#,##0.00'
                    _cell = sheet.cell(raw_range + 3, 3)
                    _cell.number_format = '#,##0.00'
                    _cell = sheet.cell(raw_range + 3, 4)
                    _cell.number_format = '#,##0.00'
                    cell_title = sheet.cell(raw_range + 3, col_range)
                    cell_title.fill = PatternFill(
                        start_color="00eaeaea", end_color="00eaeaea", fill_type="solid")
                    cell_title = sheet.cell(raw_range + 3, col_range)
                    cell_title.font = Font(bold=True, size=11)
                    bd = Side(style='thin', color="000000")
                    cell_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    def write_bookings_to_excel_file(self, booking_path, shift):
        filepath = os.path.join(booking_path, 'bookings.xlsx')
        print('Writing to excel ...')

        # if os.path.exists(filepath) and self.fork != 1:
        #     remove(filepath)
        #     self.fork = 1
        if not os.path.exists(filepath):
            workbook = Workbook()
            workbook.save(filepath)
            workbook.create_sheet("Spapshoot", 0)
            workbook.create_sheet("Display", 1)
        else:
            workbook = load_workbook(filepath)

        # fixme: delete the default sheet:
        if "Sheet" in workbook.sheetnames:
            std = workbook["Sheet"]
            workbook.remove(std)
        sheet = workbook.active
        self.set_stylesheet(sheet, self.shift)
        if shift != 1:
            # write snap sheet
            c = '1.374'
            i = 2
            for row in self.data:
                cell_reference = sheet.cell(row=i, column=1)
                cell_reference.value = row[2]
                sheet['B{}'.format(i)] = '=PRODUCT(A{},{}'.format(i, c)
                sheet['C{}'.format(i)] = '=SUM(B{},-A{}'.format(i, i)
                cell_reference = sheet.cell(row=i, column=4)
                cell_reference.value = row[0]
                cell_reference = sheet.cell(row=i, column=5)
                cell_reference.value = row[1]
                cell_reference = sheet.cell(row=i, column=6)
                cell_reference.value = row[3]
                i += 1
        else:
            # write turbo sheet
            c = '1.374'
            i = 3
            for row in self.data:
                cell_reference = sheet.cell(row=i, column=1)
                update_code = t.code_builder(self.read_code())
                self.write_code(update_code)
                cell_reference.value = update_code
                cell_reference = sheet.cell(row=i, column=2)
                cell_reference.value = row[2]
                # REF:
                # https://stackoverflow.com/questions/51044736/openpyxl-iterate-through-rows-and-apply-formula
                # fixme CODE:
                #  for row_num in range(2, max_row_num):
                #     sheet['E{}'.format(row_num)] = '=CLEAN(D{})'.format(row_num)
                sheet['C{}'.format(i)] = '=PRODUCT(B{},{}'.format(i, c)
                sheet['D{}'.format(i)] = '=SUM(C{},-B{}'.format(i, i)
                cell_reference = sheet.cell(row=i, column=5)
                cell_reference.value = self.cell_cc
                cell_reference = sheet.cell(row=i, column=6)
                cell_reference.value = self.cell_city
                cell_reference = sheet.cell(row=i, column=7)
                cell_reference.value = row[0]
                cell_reference = sheet.cell(row=i, column=8)
                cell_reference.value = row[1]
                cell_reference = sheet.cell(row=i, column=9)
                cell_reference.value = row[3]
                i += 1

            workbook.active = 1
            display_sheet = workbook.active

            # select target row
            # target = 1
            # while sheet.cell(row=target, column=1).value is not None:
            #     target += 1
            c = '1.374'
            target = 3
            while display_sheet.cell(row=target, column=6).value is not None:
                target += 1

            booking = self.data[0]
            cell_reference = display_sheet.cell(row=target, column=1)
            update_code = t.code_builder(self.read_code())
            self.write_code(update_code)
            cell_reference.value = update_code
            cell_reference = display_sheet.cell(row=target, column=2)
            cell_reference.value = booking[2]
            display_sheet['C{}'.format(target)] = '=PRODUCT(B{},{}'.format(target, c)
            display_sheet['D{}'.format(target)] = '=SUM(C{},-B{}'.format(target, target)
            cell_reference = display_sheet.cell(row=target, column=5)
            cell_reference.value = self.cell_cc
            cell_reference = display_sheet.cell(row=target, column=6)
            cell_reference.value = self.cell_city
            cell_reference = display_sheet.cell(row=target, column=7)
            cell_reference.value = booking[0]
            cell_reference = display_sheet.cell(row=target, column=8)
            cell_reference.value = booking[1]
            cell_reference = display_sheet.cell(row=target, column=9)
            cell_reference.value = booking[3]

            # switch sheet
            workbook.active = 0

            sheet = workbook.active
            self.set_stylesheet(sheet, 1)
            self.set_stylesheet(display_sheet, 1)

        workbook.save(filepath)  # save file

        if switch != 1:
            spreadsheet = '//home/jmartorell/Booking/bookings/bookings.xlsx'
            self.send_attachment(spreadsheet)

    def send_attachment(self, file):
        subject = "An email with attachment from Python"
        body = "This is an email with attachment sent from Python"
        sender_email = "jetro4100@gmail.com"
        receiver_email = "martorelljosep@gmail.com"
        # password = input("Type your password and press enter:")
        password = 'ZXspectrum5128$}_'

        # Create a multipart message and set headers
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Bcc"] = receiver_email  # Recommended for mass emails

        # Add body to email
        message.attach(MIMEText(body, "plain"))

        filename = file  # In same directory as script

        # Open PDF file in binary mode
        with open(filename, "rb") as attachment:
            # Add file as application/octet-stream
            # Email client can usually download this automatically as attachment
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        # Encode file in ASCII characters to send by email
        encoders.encode_base64(part)

        # Add header as key/value pair to attachment part
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )

        # Add attachment to message and convert message to string
        message.attach(part)
        text = message.as_string()

        # Log in to server using secure context and send email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)
            print('Sending email ...')


if __name__ == '__main__':
    switch = t.switch
    if switch != 0:
        x = 0
        while x < 25:
            app = App(depart_m=t.depart_month,
                      depart_w=t.depart_week,
                      depart_d=t.depart_day,
                      return_m=t.return_month,
                      return_w=t.return_week,
                      return_d=t.return_day,
                      target_city=d.tour_en[x][0],
                      cell_city=d.tour_en[x][0],
                      cell_cc=d.tour_en[x][1]
                      )
            x += 1
    else:
        app = App(depart_m=t.depart_month,
                  depart_w=t.depart_week,
                  depart_d=t.depart_day,
                  return_m=t.return_month,
                  return_w=t.return_week,
                  return_d=t.return_day,
                  )

